import tkinter as tk
from tkinter import ttk, filedialog, messagebox
import tkinter.font as tkfont
import webbrowser
import urllib.parse
import os
import sys
import ctypes
import re
import configparser
from datetime import datetime, date
from typing import Any, Dict, List, Tuple, Optional

try:
    import openpyxl
except ImportError as e:
    raise SystemExit("openpyxl が必要です: pip install openpyxl") from e


# =====================
# Windows AppUserModelID
# =====================
def set_appusermodel_id(app_id: str = "Ame.AISearchViewerLite") -> None:
    try:
        ctypes.windll.shell32.SetCurrentProcessExplicitAppUserModelID(app_id)
    except Exception:
        pass


set_appusermodel_id()


def app_dir() -> str:
    """設定ファイル(search_engines.ini, config.ini)を置く場所。
    PyInstaller --onefile の場合は exe のあるフォルダ。
    通常実行はこの .py のあるフォルダ。
    """
    if getattr(sys, "frozen", False):
        return os.path.dirname(sys.executable)
    return os.path.dirname(os.path.abspath(__file__))


def resource_path(relative_path: str) -> str:
    """PyInstaller 同梱リソース（ICO等）取得用"""
    try:
        base_path = sys._MEIPASS  # type: ignore[attr-defined]
    except Exception:
        base_path = app_dir()
    return os.path.join(base_path, relative_path)


# =====================
# Utilities
# =====================
def safe_str(v: Any) -> str:
    if v is None:
        return ""
    if isinstance(v, (datetime, date)):
        return v.isoformat(sep=" ")
    return str(v)


def normalize_query(s: str) -> str:
    if not s:
        return ""
    s = s.replace("\r", " ").replace("\n", " ")
    s = s.replace("　", " ")
    s = re.sub(r"\s+", " ", s)
    return s.strip()


def truncate(s: str, n: int = 80) -> str:
    s = s or ""
    return s[:n] + ("…" if len(s) > n else "")


DEFAULT_ENGINES: Dict[str, str] = {
    "Google": "https://www.google.com/search?q={query}",
    "Bing": "https://www.bing.com/search?q={query}",
    "DuckDuckGo": "https://duckduckgo.com/?q={query}",
    "Perplexity": "https://www.perplexity.ai/search?q={query}",
}


def ensure_default_files() -> Tuple[str, str]:
    """search_engines.ini / config.ini が無ければ作る"""
    d = app_dir()
    engines_path = os.path.join(d, "search_engines.ini")
    config_path = os.path.join(d, "config.ini")

    if not os.path.exists(engines_path):
        cfg = configparser.ConfigParser()
        for key, url in DEFAULT_ENGINES.items():
            sec = key.lower()
            cfg[sec] = {"name": key, "url": url}
        with open(engines_path, "w", encoding="utf-8") as f:
            cfg.write(f)

    if not os.path.exists(config_path):
        cfg = configparser.ConfigParser()
        cfg["general"] = {"default_engine": "Google", "alt_engine": "Perplexity"}
        with open(config_path, "w", encoding="utf-8") as f:
            cfg.write(f)

    return engines_path, config_path


def load_engines(path: str) -> Dict[str, str]:
    cfg = configparser.ConfigParser()
    cfg.read(path, encoding="utf-8")
    engines: Dict[str, str] = {}

    for sec in cfg.sections():
        name = cfg.get(sec, "name", fallback=sec)
        url = cfg.get(sec, "url", fallback="")
        if "{query}" not in url:
            continue
        engines[name] = url

    if not engines:
        engines = DEFAULT_ENGINES.copy()
    return engines


def load_config(path: str) -> configparser.ConfigParser:
    cfg = configparser.ConfigParser()
    cfg.read(path, encoding="utf-8")
    if "general" not in cfg:
        cfg["general"] = {}
    return cfg


def save_config(cfg: configparser.ConfigParser, path: str) -> None:
    try:
        with open(path, "w", encoding="utf-8") as f:
            cfg.write(f)
    except Exception:
        pass


class XlsxSearchViewer(tk.Tk):
    def __init__(self) -> None:
        super().__init__()

        # files
        self.engines_path, self.config_path = ensure_default_files()
        self.cfg = load_config(self.config_path)
        self.engines = load_engines(self.engines_path)

        self.title("AISearchViewerLite")
        self.geometry("1120x720")
        self.minsize(900, 520)

        try:
            self.iconbitmap(resource_path("AISearchViewer.ico"))
        except Exception:
            pass

        # workbook
        self.file_path: str = ""
        self.wb = None
        self.sheet_names: List[str] = []

        # data model
        self.headers: List[str] = []
        self.rows_all: List[List[str]] = []
        self.rows_view: List[List[str]] = []

        # selection
        self._rc_col_index: Optional[int] = None

        # sort state
        self._sort_state: Dict[int, bool] = {}  # col_index -> asc(True)/desc(False)

        # cell highlight window (border-only)
        self._hl_enabled = os.name == "nt"
        self._hl_win: Optional[tk.Toplevel] = None
        self._hl_keycolor = "#ff00ff"  # transparent key color

        self._build_ui()
        self._apply_config_to_ui()
        self._build_cell_highlight_window()

    # ---------------- UI ----------------
    def _build_ui(self) -> None:
        # Treeview selection の青を消して目立たなくする
        style = ttk.Style()
        style.map(
            "Treeview",
            background=[("selected", "white")],
            foreground=[("selected", "black")],
        )

        top = ttk.Frame(self, padding=(10, 10, 10, 6))
        top.pack(fill="x")

        ttk.Button(top, text="開く", command=self.open_file).pack(side="left")

        ttk.Label(top, text="シート:").pack(side="left", padx=(12, 6))
        self.sheet_var = tk.StringVar(value="")
        self.sheet_combo = ttk.Combobox(top, textvariable=self.sheet_var, state="readonly", width=26)
        self.sheet_combo.pack(side="left")
        self.sheet_combo.bind("<<ComboboxSelected>>", lambda e: self.load_sheet())

        ttk.Label(top, text="フィルタ:").pack(side="left", padx=(14, 6))
        self.filter_var = tk.StringVar(value="")
        self.entry_filter = ttk.Entry(top, textvariable=self.filter_var, width=36)
        self.entry_filter.pack(side="left")
        self.entry_filter.bind("<KeyRelease>", lambda e: self.apply_filter())
        self.entry_filter.bind("<Return>", lambda e: self.search_default_engine())

        ttk.Button(top, text="クリア", command=self.clear_filter).pack(side="left", padx=(8, 0))

        center = ttk.Frame(self, padding=(10, 0, 10, 0))
        center.pack(fill="both", expand=True)

        self.tree = ttk.Treeview(center, show="headings", selectmode="browse")
        self.tree.grid(row=0, column=0, sticky="nsew")

        self.vsb = ttk.Scrollbar(center, orient="vertical", command=self._on_yscroll)
        self.hsb = ttk.Scrollbar(center, orient="horizontal", command=self._on_xscroll)
        self.vsb.grid(row=0, column=1, sticky="ns")
        self.hsb.grid(row=1, column=0, sticky="ew")
        self.tree.configure(yscrollcommand=self.vsb.set, xscrollcommand=self.hsb.set)

        center.grid_rowconfigure(0, weight=1)
        center.grid_columnconfigure(0, weight=1)

        # Zebra tags
        self.tree.tag_configure("odd", background="white")
        self.tree.tag_configure("even", background="#f3f3f3")

        # Bindings
        self.tree.bind("<Button-3>", self.on_right_click)
        self.tree.bind("<<TreeviewSelect>>", lambda e: self._after_select())
        self.tree.bind("<Double-1>", lambda e: self.show_full_text())

        # main window resize -> update highlight
        self.bind("<Configure>", lambda e: self.draw_cell_highlight())

        bottom = ttk.Frame(self, padding=(10, 6, 10, 10))
        bottom.pack(fill="x")

        self.status_var = tk.StringVar(value="ファイル未読込")
        ttk.Label(bottom, textvariable=self.status_var, anchor="w").pack(side="left", fill="x", expand=True)

        self.engine_var = tk.StringVar(value="Google")
        ttk.Label(bottom, text="Enter検索:").pack(side="left", padx=(10, 6))
        self.engine_combo = ttk.Combobox(bottom, textvariable=self.engine_var, state="readonly",
                                         width=14, values=list(self.engines.keys()))
        self.engine_combo.pack(side="left")
        self.engine_combo.bind("<<ComboboxSelected>>", lambda e: self.on_engine_changed())

        # Right-click menu
        self.menu = tk.Menu(self, tearoff=0)
        self.menu_search = tk.Menu(self.menu, tearoff=0)
        self.menu.add_cascade(label="検索", menu=self.menu_search)
        self.menu.add_separator()
        self.menu.add_command(label="全文表示", command=self.show_full_text)
        self.menu.add_command(label="コピー", command=self.copy_cell_text)
        self.menu.add_command(label="検索URLをコピー", command=self.copy_search_url)

        # keyboard shortcuts
        self.bind_all("<Control-c>", lambda e: self.copy_cell_text())
        self.bind_all("<Control-Shift-C>", lambda e: self.copy_search_url())
        self.bind_all("<Control-L>", lambda e: self.focus_filter())
        self.bind_all("<Control-Return>", lambda e: self.search_alt_engine())
        self.bind_all("<F2>", lambda e: self.show_full_text())

    def _apply_config_to_ui(self) -> None:
        default_engine = self.cfg.get("general", "default_engine", fallback="Google")
        if default_engine in self.engines:
            self.engine_var.set(default_engine)
        else:
            self.engine_var.set(next(iter(self.engines.keys())))
        self.engine_combo["values"] = list(self.engines.keys())

    def focus_filter(self) -> None:
        self.entry_filter.focus_set()
        self.entry_filter.select_range(0, tk.END)

    # ---------------- Scroll hooks (枠のズレ防止) ----------------
    def _on_yscroll(self, *args) -> None:
        self.tree.yview(*args)
        self.draw_cell_highlight()

    def _on_xscroll(self, *args) -> None:
        self.tree.xview(*args)
        self.draw_cell_highlight()

    # ---------------- Config ----------------
    def on_engine_changed(self) -> None:
        self.cfg["general"]["default_engine"] = self.engine_var.get()
        save_config(self.cfg, self.config_path)
        self.update_status(extra="既定エンジンを保存")

    def _alt_engine(self) -> str:
        alt = self.cfg.get("general", "alt_engine", fallback="Perplexity")
        if alt in self.engines:
            return alt
        return next(iter(self.engines.keys()))

    # ---------------- File / Sheet ----------------
    def open_file(self) -> None:
        fp = filedialog.askopenfilename(
            title="Excelファイルを選択",
            filetypes=[("Excel", "*.xlsx *.xlsm"), ("All files", "*.*")]
        )
        if not fp:
            return
        try:
            self.wb = openpyxl.load_workbook(fp, read_only=True, data_only=False)
        except Exception as e:
            messagebox.showerror("読み込み失敗", f"{e}")
            return

        self.file_path = fp
        self.sheet_names = self.wb.sheetnames
        self.sheet_combo["values"] = self.sheet_names
        self.sheet_var.set(self.sheet_names[0] if self.sheet_names else "")
        self.load_sheet()

    def load_sheet(self) -> None:
        if not self.wb:
            return
        ws = self.wb[self.sheet_var.get()]

        max_row = ws.max_row or 0
        max_col = ws.max_column or 0

        values: List[List[str]] = []
        for row in ws.iter_rows(min_row=1, max_row=max_row, max_col=max_col, values_only=False):
            values.append([safe_str(cell.value) for cell in row])

        # trim trailing empty columns
        def col_empty(ci: int) -> bool:
            for r in values:
                if ci < len(r) and r[ci] != "":
                    return False
            return True

        last_col = max_col
        while last_col > 0 and col_empty(last_col - 1):
            last_col -= 1
        values = [r[:last_col] for r in values]

        col_letters = [openpyxl.utils.get_column_letter(i) for i in range(1, last_col + 1)]
        self.headers = ["#"] + col_letters

        self.rows_all = [[str(i)] + r for i, r in enumerate(values, start=1)]
        self.rows_view = self.rows_all[:]
        self.filter_var.set("")
        self._sort_state.clear()
        self._render_table()
        self.update_status()
        self.hide_cell_highlight()

    # ---------------- Rendering ----------------
    def _render_table(self) -> None:
        self.tree.delete(*self.tree.get_children())

        self.tree["columns"] = list(range(len(self.headers)))
        for idx, h in enumerate(self.headers):
            self.tree.heading(idx, text=h, command=lambda c=idx: self.sort_by_column(c))
            w = 60 if idx == 0 else 140
            self.tree.column(idx, width=w, minwidth=50, stretch=True, anchor="w")

        for i, row in enumerate(self.rows_view):
            tag = "even" if i % 2 == 0 else "odd"
            self.tree.insert("", "end", values=row, tags=(tag,))

        # 列幅自動調整（列ごと）
        self.autosize_columns(sample_rows=300)

    def autosize_columns(self, sample_rows: int = 300, padding: int = 18, max_width: int = 520) -> None:
        if not self.headers:
            return
        font = tkfont.nametofont("TkDefaultFont")
        widths = [font.measure(h) + padding for h in self.headers]

        for row in self.rows_view[:sample_rows]:
            for i, val in enumerate(row):
                w = font.measure(str(val)) + padding
                if w > widths[i]:
                    widths[i] = w

        for i, w in enumerate(widths):
            if i == 0:
                w = min(max(w, 60), 80)
            else:
                w = min(max(w, 60), max_width)
            self.tree.column(i, width=w)

    # ---------------- Sort ----------------
    def sort_by_column(self, col_index: int) -> None:
        if not self.rows_view:
            return

        asc = self._sort_state.get(col_index, True)
        self._sort_state[col_index] = not asc

        if col_index == 0:
            def key_func(r: List[str]) -> int:
                try:
                    return int(r[0])
                except Exception:
                    return 0
        else:
            def key_func(r: List[str]) -> str:
                try:
                    return (r[col_index] or "").lower()
                except Exception:
                    return ""

        try:
            self.rows_view.sort(key=key_func, reverse=not asc)
        except Exception:
            return

        self._render_table()
        self.update_status(extra=f"{self.headers[col_index]} で{'昇順' if asc else '降順'}ソート")
        self.draw_cell_highlight()

    # ---------------- Filter ----------------
    def clear_filter(self) -> None:
        self.filter_var.set("")
        self.apply_filter()

    def apply_filter(self) -> None:
        q = self.filter_var.get().strip().lower()
        if not q:
            self.rows_view = self.rows_all[:]
        else:
            self.rows_view = [r for r in self.rows_all if q in "\t".join(r).lower()]
        self._render_table()
        self.update_status()
        self.draw_cell_highlight()

    # ---------------- Selection helpers ----------------
    def _get_selected_values(self) -> List[str]:
        sel = self.tree.selection()
        if not sel:
            return []
        return list(self.tree.item(sel[0], "values") or [])

    def _get_selected_cell_text_raw(self) -> str:
        vals = self._get_selected_values()
        if not vals:
            return ""
        ci = self._rc_col_index if self._rc_col_index is not None else 1
        if 0 <= ci < len(vals):
            return str(vals[ci])
        return ""

    def _get_selected_cell_text(self) -> str:
        return normalize_query(self._get_selected_cell_text_raw())

    def _after_select(self) -> None:
        self.update_status()
        self.draw_cell_highlight()

    # ---------------- Cell highlight (枠だけ) ----------------
    def _build_cell_highlight_window(self) -> None:
        if not self._hl_enabled:
            return

        win = tk.Toplevel(self)
        win.overrideredirect(True)
        win.attributes("-topmost", True)
        win.configure(bg=self._hl_keycolor)
        try:
            # Windows: 指定色を透明にできる
            win.wm_attributes("-transparentcolor", self._hl_keycolor)
        except Exception:
            # 透明化できない環境では枠表示を無効にする
            win.destroy()
            self._hl_enabled = False
            self._hl_win = None
            return

        # 枠だけ（中は透明色＝透過）
        frm = tk.Frame(
            win,
            bg=self._hl_keycolor,
            highlightthickness=2,
            highlightbackground="#2b6cff",
            highlightcolor="#2b6cff"
        )
        frm.pack(fill="both", expand=True)

        win.withdraw()
        self._hl_win = win

    def hide_cell_highlight(self) -> None:
        if self._hl_win is not None:
            try:
                self._hl_win.withdraw()
            except Exception:
                pass

    def draw_cell_highlight(self) -> None:
        if not self._hl_enabled or self._hl_win is None:
            return

        sel = self.tree.selection()
        if not sel or self._rc_col_index is None:
            self.hide_cell_highlight()
            return

        item = sel[0]
        col = f"#{self._rc_col_index + 1}"
        bbox = self.tree.bbox(item, col)  # (x, y, w, h)
        if not bbox:
            self.hide_cell_highlight()
            return

        x, y, w, h = bbox
        rx = self.tree.winfo_rootx() + x
        ry = self.tree.winfo_rooty() + y

        # 枠が見やすいように少しだけ外へ
        pad = 0
        geom = f"{max(1, w + pad)}x{max(1, h + pad)}+{rx}+{ry}"

        try:
            self._hl_win.geometry(geom)
            self._hl_win.deiconify()
        except Exception:
            pass

    # ---------------- Right-click / Menu ----------------
    def on_right_click(self, event) -> None:
        row_id = self.tree.identify_row(event.y)
        col_id = self.tree.identify_column(event.x)
        if not row_id or not col_id:
            return

        self.tree.selection_set(row_id)
        try:
            self._rc_col_index = int(col_id[1:]) - 1
        except Exception:
            self._rc_col_index = 1

        self.draw_cell_highlight()
        self._rebuild_engine_menu()

        try:
            self.menu.tk_popup(event.x_root, event.y_root)
        finally:
            self.menu.grab_release()

    def _rebuild_engine_menu(self) -> None:
        self.menu_search.delete(0, "end")

        preview = self._get_selected_cell_text()
        if preview:
            self.menu_search.add_command(label=f"検索語句: {truncate(preview, 60)}", state="disabled")
            self.menu_search.add_separator()

        for name in self.engines.keys():
            self.menu_search.add_command(label=f"{name}で検索", command=lambda n=name: self.search_with_engine(n))

        self.engine_combo["values"] = list(self.engines.keys())

    # ---------------- Search ----------------
    def _make_search_url(self, engine_name: str, query: str) -> str:
        tpl = self.engines.get(engine_name, DEFAULT_ENGINES["Google"])
        enc = urllib.parse.quote(query, safe="")
        return tpl.replace("{query}", enc)

    def search_with_engine(self, engine_name: str) -> None:
        q = self._get_selected_cell_text()
        if not q:
            messagebox.showinfo("検索", "空のセルは検索できません。")
            return
        webbrowser.open(self._make_search_url(engine_name, q))
        self.update_status(extra=f"{engine_name}で検索")

    def search_default_engine(self) -> None:
        q = self._get_selected_cell_text()
        if not q:
            return
        eng = self.engine_var.get()
        webbrowser.open(self._make_search_url(eng, q))
        self.update_status(extra=f"{eng}で検索")

    def search_alt_engine(self) -> None:
        q = self._get_selected_cell_text()
        if not q:
            return
        eng = self._alt_engine()
        webbrowser.open(self._make_search_url(eng, q))
        self.update_status(extra=f"{eng}で検索(Ctrl+Enter)")

    # ---------------- Copy / Full text ----------------
    def copy_cell_text(self) -> None:
        text = self._get_selected_cell_text_raw()
        if text == "":
            return
        self.clipboard_clear()
        self.clipboard_append(text)
        self.update_status(extra="コピーしました")

    def copy_search_url(self) -> None:
        text = self._get_selected_cell_text()
        if text == "":
            return
        url = self._make_search_url(self.engine_var.get(), text)
        self.clipboard_clear()
        self.clipboard_append(url)
        self.update_status(extra="URLをコピーしました")

    def show_full_text(self) -> None:
        raw = self._get_selected_cell_text_raw()
        if raw == "":
            return

        win = tk.Toplevel(self)
        win.title("全文表示")
        win.geometry("720x450")
        win.transient(self)
        win.grab_set()

        frm = ttk.Frame(win, padding=10)
        frm.pack(fill="both", expand=True)

        ttk.Label(frm, text="選択セルの内容（表示そのまま）").pack(anchor="w")

        txt = tk.Text(frm, wrap="word")
        txt.pack(fill="both", expand=True, pady=(8, 10))
        txt.insert("1.0", raw)
        txt.configure(state="disabled")

        btns = ttk.Frame(frm)
        btns.pack(fill="x")
        ttk.Button(btns, text="コピー", command=lambda: self._copy_from_popup(raw)).pack(side="left")
        ttk.Button(btns, text="閉じる", command=win.destroy).pack(side="right")

    def _copy_from_popup(self, s: str) -> None:
        self.clipboard_clear()
        self.clipboard_append(s)
        self.update_status(extra="コピーしました")

    # ---------------- Status ----------------
    def update_status(self, extra: str = "") -> None:
        fname = os.path.basename(self.file_path) if self.file_path else "(未読込)"
        sheet = self.sheet_var.get() or "-"
        total = len(self.rows_all)
        shown = len(self.rows_view)

        cell_info = ""
        vals = self._get_selected_values()
        if vals:
            ci = self._rc_col_index if self._rc_col_index is not None else 1
            col_name = self.headers[ci] if 0 <= ci < len(self.headers) else "?"
            row_no = vals[0] if vals else "?"
            q = self._get_selected_cell_text()
            cell_info = f"  選択: R{row_no} {col_name}  {len(q)}文字  プレビュー: {truncate(q, 80)}"

        msg = f"{fname} / {sheet}   行: {shown}/{total}{cell_info}"
        if extra:
            msg += f"   | {extra}"
        self.status_var.set(msg)


def main() -> None:
    app = XlsxSearchViewer()
    app.mainloop()


if __name__ == "__main__":
    main()
